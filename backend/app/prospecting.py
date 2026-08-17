import hashlib
import csv
import io
import os
import re
import sys
from urllib.parse import urljoin, urlparse
from datetime import datetime, timezone
from pathlib import Path

import httpx
from dotenv import load_dotenv
from sqlalchemy import select

from .database import SessionLocal
from .models import Prospect, ProspectingJob
from .lead_quality import assess_lead
from .prospect_messages import generate_prospect_messages


BACKEND_ROOT = Path(__file__).resolve().parents[1]
load_dotenv(BACKEND_ROOT / ".env")
EXECUTION = BACKEND_ROOT / "execution"
if str(EXECUTION) not in sys.path:
    sys.path.insert(0, str(EXECUTION))


def _email_from_site(url: str) -> str:
    if not url:
        return ""
    try:
        response = httpx.get(url, follow_redirects=True, timeout=8, headers={"User-Agent": "Mozilla/5.0"})
        if response.status_code >= 400:
            return ""
        pages = [(url, response.text)]
        links = re.findall(r'href=["\']([^"\']+)["\']', response.text, re.I)
        keywords = ("contact", "mentions", "legal", "about", "a-propos", "equipe")
        host = urlparse(str(response.url)).netloc
        for link in links:
            absolute = urljoin(str(response.url), link)
            if urlparse(absolute).netloc == host and any(word in absolute.lower() for word in keywords):
                try:
                    child = httpx.get(absolute, follow_redirects=True, timeout=6, headers={"User-Agent": "Mozilla/5.0"})
                    if child.status_code < 400: pages.append((absolute, child.text))
                except Exception: pass
                if len(pages) >= 4: break
        matches = []
        for _, html in pages:
            matches.extend(re.findall(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", html, re.I))
        rejected = ("example.", "wixpress", "sentry", "webpack", "domain.com")
        return next((mail.lower() for mail in matches if not any(bad in mail.lower() for bad in rejected)), "")
    except Exception:
        return ""


def run_prospecting_job(job_id: int) -> None:
    from scrape_google_maps import scrape_google_maps
    from enrich_dirigeants import enrich_lead
    with SessionLocal() as db:
        job = db.get(ProspectingJob, job_id)
        if not job:
            return
        job.status = "running"; db.commit()
        try:
            businesses = scrape_google_maps(job.query, job.requested_limit, job.location, language="fr")
            scraped_emails = 0
            for item in businesses:
                name = item.get("title") or ""
                address = item.get("address") or ""
                place_id = item.get("placeId") or ""
                key = hashlib.sha256(f"{place_id or name}|{address}".lower().encode()).hexdigest()
                existing = db.scalar(select(Prospect).where(Prospect.dedupe_key == key))
                email = item.get("email") or ""
                source = "google_maps" if email else ""
                if not email:
                    email = _email_from_site(item.get("website") or "")
                    source = "website" if email else ""
                values = dict(job_id=job.id, business_name=name, category=item.get("categoryName") or "", address=address, city=item.get("city") or "", country=item.get("countryCode") or "", phone=item.get("phone") or "", email=email, email_source=source, website=item.get("website") or "", google_maps_url=item.get("url") or "", place_id=place_id, rating=str(item.get("totalScore") or ""), review_count=int(item.get("reviewsCount") or 0), raw_data=item)
                if values["country"].upper() == "FR" or re.search(r"\b\d{5}\b", address):
                    try:
                        enrichment = enrich_lead(name, zip_code=(re.search(r"\b\d{5}\b", address).group(0) if re.search(r"\b\d{5}\b", address) else None), website=values["website"], gmaps_category=values["category"])
                        values["decision_maker_name"] = " ".join(filter(None, (enrichment.get("dirigeant_prenom"), enrichment.get("dirigeant_nom"))))
                        values["decision_maker_role"] = enrichment.get("dirigeant_qualite", "")
                    except Exception:
                        pass
                quality = assess_lead(values)
                values.update(quality)
                values["score"] = quality["quality_score"]
                if values["email"]: scraped_emails += 1
                if quality["validation_status"] != "qualified":
                    continue
                if quality["validation_status"] == "qualified":
                    try:
                        messages = generate_prospect_messages(values)
                        if messages: values.update(messages.model_dump())
                    except Exception as message_error:
                        values["validation_notes"] += f"; rédaction IA indisponible: {message_error}"
                if existing:
                    for field, value in values.items():
                        if value not in ("", None, {}): setattr(existing, field, value)
                else:
                    db.add(Prospect(dedupe_key=key, **values))
            job.status = "completed"; job.found_count = len(businesses); job.email_scraped_count = scraped_emails; job.completed_at = datetime.now(timezone.utc)
        except Exception as exc:
            job.status = "failed"; job.error = str(exc); job.completed_at = datetime.now(timezone.utc)
        db.commit()


def import_prospect_file(filename: str, content: bytes) -> dict:
    rows = []
    if filename.lower().endswith(".csv"):
        text = content.decode("utf-8-sig")
        try: dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
        except csv.Error: dialect = csv.excel
        rows = list(csv.DictReader(io.StringIO(text), dialect=dialect))
    elif filename.lower().endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        book = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = book.active
        values = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(values, [])]
        rows = [dict(zip(headers, [str(value or "") for value in row])) for row in values]
    else:
        raise ValueError("Formats acceptés : CSV, XLSX et XLSM")
    aliases = {"business_name": ("business_name","entreprise","company","nom_raison_sociale"), "category": ("category","categorie","secteur","activité","activite"), "address": ("address","adresse"), "city": ("city","ville"), "country": ("country","pays"), "phone": ("phone","telephone","téléphone","mobile","whatsapp"), "email": ("email","mail","email professionnel"), "website": ("website","site","site_web","url"), "decision_maker_name": ("decision_maker_name","decideur","dirigeant","nom_complet"), "decision_maker_role": ("decision_maker_role","role","fonction","poste"), "decision_maker_linkedin": ("decision_maker_linkedin","linkedin","linkedin_url"), "description": ("description","présentation","presentation","activité détaillée","activite detaillee"), "tags": ("tags","segments","labels"), "notes": ("notes","commentaires","commentaire"), "email_subject": ("email_subject","objet_email","objet email","sujet email"), "email_message": ("email_message","message_email","message email","email pré-écrit","email pre-ecrit"), "whatsapp_message": ("whatsapp_message","message_whatsapp","message whatsapp","whatsapp pré-écrit","whatsapp pre-ecrit"), "sms_message": ("sms_message","message_sms","message sms","sms pré-écrit","sms pre-ecrit")}
    added = updated = invalid = emails = 0
    with SessionLocal() as db:
        for source in rows:
            low = {str(k).strip().lower(): str(v or "").strip() for k,v in source.items() if k}
            mapped = {field: next((low[a] for a in names if low.get(a)), "") for field,names in aliases.items()}
            if not mapped["decision_maker_name"]:
                mapped["decision_maker_name"] = " ".join(value for value in (low.get("prenom") or low.get("prénom") or low.get("first_name"), low.get("nom") or low.get("last_name")) if value).strip()
            if not mapped["business_name"] and not mapped["email"] and not mapped["phone"]:
                invalid += 1; continue
            email = mapped["email"].lower()
            key = hashlib.sha256(f"{email or mapped['phone'] or mapped['business_name']}|{mapped['address']}".lower().encode()).hexdigest()
            existing = db.scalar(select(Prospect).where(Prospect.dedupe_key == key))
            values = {**mapped, "email": email, "email_source": "file" if email else "", "raw_data": low}
            quality = assess_lead(values)
            values.update(quality)
            values["score"] = quality["quality_score"]
            if len(values.get("sms_message", "")) > 160:
                values["sms_message"] = values["sms_message"][:160]
            if quality["validation_status"] == "qualified" and not any((values.get("email_message"), values.get("whatsapp_message"), values.get("sms_message"))):
                try:
                    messages = generate_prospect_messages(values)
                    if messages: values.update(messages.model_dump())
                except Exception as message_error:
                    values["validation_notes"] += f"; rédaction IA indisponible: {message_error}"
            if values["email"]: emails += 1
            if quality["validation_status"] == "rejected": invalid += 1
            if quality["validation_status"] != "qualified":
                continue
            if existing:
                for field,value in values.items():
                    if value not in ("",None,{}): setattr(existing,field,value)
                updated += 1
            else:
                db.add(Prospect(dedupe_key=key, **values)); added += 1
        db.commit()
    return {"added":added,"updated":updated,"invalid":invalid,"emails":emails}
